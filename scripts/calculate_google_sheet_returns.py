import argparse
import csv
import json
import sqlite3
import sys
from urllib.parse import unquote
from datetime import date, datetime, timezone
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from scripts.calculate_sp500_benchmark import (
    compound_annual_return as benchmark_compound_annual_return,
    load_sp500_series,
    value_at_or_before,
)
from scripts.calculate_total_returns import (
    DEFAULT_QUICKFS_DB,
    DEFAULT_VIC_DB,
    calculate_return,
    compound_annual_return,
    find_series,
    idea_month,
    load_quickfs_series,
    normalize_ticker,
    start_index,
)


DEFAULT_INPUT = Path("/private/tmp/vic_input_ideas.xlsx")
DEFAULT_OUTPUT_JSON = Path("analysis/google_sheet_idea_returns.json")
DEFAULT_OUTPUT_CSV = Path("analysis/google_sheet_idea_returns.csv")


RESULT_FIELDS = [
    "source_sheet",
    "source_row",
    "ticker",
    "sheet_date",
    "match_status",
    "direction",
    "vic_db_direction",
    "company_name",
    "db_company_name",
    "match_key",
    "idea_id",
    "idea_link",
    "is_contest_winner",
    "matched_ticker",
    "start_period",
    "end_period",
    "start_price",
    "end_price",
    "dividends",
    "years_held",
    "stock_total_return_pct",
    "idea_total_return_pct",
    "annualized_idea_return_pct",
    "benchmark_total_return_pct",
    "benchmark_annualized_return_pct",
    "excess_total_return_pct",
    "excess_annualized_return_pct",
]


def parse_args():
    parser = argparse.ArgumentParser(
        description="Calculate returns and S&P TR beats for ticker/date ideas in a Google Sheet export."
    )
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--vic-db", type=Path, default=DEFAULT_VIC_DB)
    parser.add_argument("--quickfs-db", type=Path, default=DEFAULT_QUICKFS_DB)
    parser.add_argument("--output-json", type=Path, default=DEFAULT_OUTPUT_JSON)
    parser.add_argument("--output-csv", type=Path, default=DEFAULT_OUTPUT_CSV)
    return parser.parse_args()


def iso_date(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text[:10], fmt).date().isoformat()
        except ValueError:
            pass
    if len(text) >= 10 and text[4] in "-/" and text[7] in "-/":
        return text[:10].replace("/", "-")
    return None


def load_sheet_rows(path):
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    rows = []
    for ws in wb.worksheets:
        for row_index in range(2, ws.max_row + 1):
            ticker = normalize_ticker(ws.cell(row=row_index, column=1).value)
            sheet_date = iso_date(ws.cell(row=row_index, column=2).value)
            if ticker and sheet_date:
                rows.append(
                    {
                        "source_sheet": ws.title,
                        "source_row": row_index,
                        "ticker": ticker,
                        "sheet_date": sheet_date,
                    }
                )
    return rows


def title_from_idea_link(link):
    if not link:
        return None
    marker = "/idea/"
    if marker not in link:
        return None
    slug = link.split(marker, 1)[1].split("/", 1)[0]
    if not slug:
        return None
    return unquote(slug).replace("_", " ").strip() or None


def load_db_matches(conn):
    matches = {}
    for row in conn.execute(
        """
        SELECT
            i.id,
            c.ticker,
            c.company_name,
            substr(i.date, 1, 10) AS idea_date,
            i.is_short,
            i.is_contest_winner,
            i.link,
            r.matched_ticker,
            r.start_period,
            r.end_period,
            r.start_price,
            r.end_price,
            r.dividends,
            r.periods_held,
            r.stock_total_return_pct,
            r.idea_total_return_pct,
            r.annualized_idea_return_pct,
            r.benchmark_total_return_pct,
            r.benchmark_annualized_return_pct,
            r.excess_total_return_pct,
            r.excess_annualized_return_pct
        FROM ideas i
        JOIN companies c ON c.ticker = i.company_id
        LEFT JOIN idea_total_returns r ON r.idea_id = i.id
        WHERE i.date IS NOT NULL
        """
    ):
        (
            idea_id,
            ticker,
            company_name,
            idea_date,
            is_short,
            is_contest_winner,
            idea_link,
            matched_ticker,
            start_period,
            end_period,
            start_price,
            end_price,
            dividends,
            periods_held,
            stock_total_return,
            idea_total_return,
            annualized_idea_return,
            benchmark_total_return,
            benchmark_annualized_return,
            excess_total_return,
            excess_annualized_return,
        ) = row
        key = (normalize_ticker(ticker), idea_date)
        matches.setdefault(key, []).append(
            {
                "idea_id": idea_id,
                "company_name": title_from_idea_link(idea_link) or company_name,
                "db_company_name": company_name,
                "match_key": f"{normalize_ticker(ticker)}|{idea_date}",
                "idea_link": idea_link,
                "vic_db_direction": "Short" if bool(is_short) else "Long",
                "direction": "Long",
                "is_contest_winner": bool(is_contest_winner),
                "matched_ticker": matched_ticker,
                "start_period": start_period,
                "end_period": end_period,
                "start_price": start_price,
                "end_price": end_price,
                "dividends": dividends,
                "years_held": None if periods_held is None else periods_held / 4,
                "stock_total_return_pct": stock_total_return,
                "idea_total_return_pct": stock_total_return,
                "annualized_idea_return_pct": compound_annual_return(
                    stock_total_return, periods_held / 4
                )
                if periods_held
                else None,
                "benchmark_total_return_pct": benchmark_total_return,
                "benchmark_annualized_return_pct": benchmark_annualized_return,
                "excess_total_return_pct": (
                    None
                    if benchmark_total_return is None or stock_total_return is None
                    else stock_total_return - benchmark_total_return
                ),
                "excess_annualized_return_pct": (
                    None
                    if stock_total_return is None
                    or benchmark_annualized_return is None
                    or not periods_held
                    or compound_annual_return(stock_total_return, periods_held / 4) is None
                    else compound_annual_return(stock_total_return, periods_held / 4)
                    - benchmark_annualized_return
                ),
            }
        )
    return matches


def choose_match(candidates):
    if not candidates:
        return None
    with_returns = [candidate for candidate in candidates if candidate["start_period"]]
    return (with_returns or candidates)[0]


def benchmark_return(periods, values, start_period, end_period, years_held):
    start_value = value_at_or_before(periods, values, start_period)
    end_value = value_at_or_before(periods, values, end_period)
    if start_value is None or end_value is None:
        return None, None
    total_return = (end_value / start_value - 1) * 100
    annual_return = benchmark_compound_annual_return(total_return, years_held)
    return total_return, annual_return


def calculate_assumed_long(row, quickfs, sp500_periods, sp500_values):
    matched_ticker, series = find_series(quickfs, row["ticker"])
    if not series:
        return {
            "match_status": "No VIC db match; no QuickFS ticker match",
            "direction": "Long",
        }

    start = start_index(series, idea_month(row["sheet_date"]))
    if start is None:
        return {
            "match_status": "No VIC db match; no QuickFS period on/after date",
            "direction": "Long",
            "matched_ticker": matched_ticker,
        }
    if start >= len(series) - 1:
        return {
            "match_status": "No VIC db match; no later QuickFS period",
            "direction": "Long",
            "matched_ticker": matched_ticker,
        }

    result = calculate_return(series, start)
    years_held = result["periods_held"] / 4
    idea_total = result["stock_total_return_pct"]
    idea_annual = compound_annual_return(idea_total, years_held)
    benchmark_total, benchmark_annual = benchmark_return(
        sp500_periods,
        sp500_values,
        result["start_period"],
        result["end_period"],
        years_held,
    )
    excess_total = None if benchmark_total is None else idea_total - benchmark_total
    excess_annual = (
        None
        if idea_annual is None or benchmark_annual is None
        else idea_annual - benchmark_annual
    )
    return {
        "match_status": "No VIC db match; calculated as long from ticker/date",
        "direction": "Long",
        "matched_ticker": matched_ticker,
        "start_period": result["start_period"],
        "end_period": result["end_period"],
        "start_price": result["start_price"],
        "end_price": result["end_price"],
        "dividends": result["dividends"],
        "years_held": years_held,
        "stock_total_return_pct": result["stock_total_return_pct"],
        "idea_total_return_pct": idea_total,
        "annualized_idea_return_pct": idea_annual,
        "benchmark_total_return_pct": benchmark_total,
        "benchmark_annualized_return_pct": benchmark_annual,
        "excess_total_return_pct": excess_total,
        "excess_annualized_return_pct": excess_annual,
    }


def summarize(rows):
    summary = {}
    for row in rows:
        sheet = row["source_sheet"]
        bucket = summary.setdefault(
            sheet,
            {
                "total_rows": 0,
                "exact_vic_matches": 0,
                "assumed_long_calculated": 0,
                "with_returns": 0,
                "avg_annual_return_pct": None,
                "avg_annual_beat_pct": None,
            },
        )
        bucket["total_rows"] += 1
        if row["match_status"].startswith("Exact VIC db match"):
            bucket["exact_vic_matches"] += 1
        if row["match_status"].startswith("No VIC db match; calculated"):
            bucket["assumed_long_calculated"] += 1
        if row.get("annualized_idea_return_pct") is not None:
            bucket["with_returns"] += 1

    for sheet, bucket in summary.items():
        sheet_rows = [row for row in rows if row["source_sheet"] == sheet]
        annuals = [
            row["annualized_idea_return_pct"]
            for row in sheet_rows
            if row.get("annualized_idea_return_pct") is not None
        ]
        beats = [
            row["excess_annualized_return_pct"]
            for row in sheet_rows
            if row.get("excess_annualized_return_pct") is not None
        ]
        if annuals:
            bucket["avg_annual_return_pct"] = sum(annuals) / len(annuals)
        if beats:
            bucket["avg_annual_beat_pct"] = sum(beats) / len(beats)
    return summary


def main():
    args = parse_args()
    sheet_rows = load_sheet_rows(args.input)
    conn = sqlite3.connect(args.vic_db)
    matches = load_db_matches(conn)
    sp500_periods, sp500_values = load_sp500_series(conn)
    conn.close()
    quickfs = load_quickfs_series(args.quickfs_db)

    results = []
    for sheet_row in sheet_rows:
        row = dict(sheet_row)
        match = choose_match(matches.get((row["ticker"], row["sheet_date"])))
        if match:
            row.update(match)
            row["match_status"] = "Exact VIC db match"
        else:
            row.update(calculate_assumed_long(row, quickfs, sp500_periods, sp500_values))
        results.append({field: row.get(field) for field in RESULT_FIELDS})

    payload = {
        "generated_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "source_workbook": str(args.input),
        "summary": summarize(results),
        "rows": results,
    }

    args.output_json.parent.mkdir(parents=True, exist_ok=True)
    args.output_json.write_text(json.dumps(payload, indent=2), encoding="utf-8")

    with args.output_csv.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=RESULT_FIELDS)
        writer.writeheader()
        writer.writerows(results)

    print(f"rows={len(results)}")
    print(f"json={args.output_json}")
    print(f"csv={args.output_csv}")
    for sheet, bucket in payload["summary"].items():
        print(
            f"{sheet}: rows={bucket['total_rows']} "
            f"exact={bucket['exact_vic_matches']} "
            f"assumed_long={bucket['assumed_long_calculated']} "
            f"with_returns={bucket['with_returns']} "
            f"avg_annual={bucket['avg_annual_return_pct']} "
            f"avg_beat={bucket['avg_annual_beat_pct']}"
        )


if __name__ == "__main__":
    main()
